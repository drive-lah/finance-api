#!/usr/bin/env python3
"""Convert ACCOUNT_VIEW_MAP.md -> ACCOUNT_VIEW_MAP.xlsx.

One row per (account x market x view). Parses the markdown pipe-tables,
tracks the current section header as the Category, colour-codes the Built?
column (green=built, grey=external-fed, yellow=not-built), autofilter + freeze.
"""
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

WIP = Path(__file__).parent
SRC = WIP / "ACCOUNT_VIEW_MAP.md"
OUT = WIP / "ACCOUNT_VIEW_MAP.xlsx"

GREEN = PatternFill("solid", fgColor="C6EFCE")
GREY = PatternFill("solid", fgColor="D9D9D9")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
CAT_FILL = PatternFill("solid", fgColor="DDEBF7")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def built_fill(status: str) -> PatternFill:
    s = status.lower()
    if s.startswith("built"):
        return GREEN
    if "external" in s:
        return GREY
    return YELLOW


def split_row(line: str):
    # strip leading/trailing pipe, split on |
    cells = [c.strip() for c in line.strip().strip("|").split("|")]
    return cells


def clean(cell: str) -> str:
    cell = cell.replace("**", "").replace("`", "")
    if cell == '"':
        return "〃"
    return cell


rows = []  # (category, coa, meaning, market, view, explanation, built)
current_cat = ""
lines = SRC.read_text().splitlines()
i = 0
while i < len(lines):
    line = lines[i]
    if line.startswith("## "):
        current_cat = line[3:].strip()
    # 3-column external-fed table: | COA | Meaning | Built? |
    if (line.strip().startswith("|") and "COA" in line and "Built?" in line
            and "Market" not in line):
        i += 2
        while i < len(lines) and lines[i].strip().startswith("|"):
            cells = [clean(c) for c in split_row(lines[i])]
            if len(cells) >= 3:
                coa, meaning, built = cells[:3]
                rows.append((current_cat, coa, meaning, "—", "(no view)", "—", built))
            i += 1
        continue
    # detect a table header row we care about
    if line.strip().startswith("|") and "COA" in line and "Built?" in line:
        header = split_row(line)
        # skip the separator row
        i += 2
        # capture last non-empty coa for 〃 fills
        last_coa = ""
        while i < len(lines) and lines[i].strip().startswith("|"):
            cells = [clean(c) for c in split_row(lines[i])]
            if len(cells) >= 6:
                coa, meaning, market, view, expl, built = cells[:6]
                if coa in ("〃", ""):
                    coa = last_coa
                else:
                    last_coa = coa
                rows.append((current_cat, coa, meaning, market, view, expl, built))
            i += 1
        continue
    i += 1

# Build workbook
wb = Workbook()
ws = wb.active
ws.title = "Account View Map"

headers = ["Category", "COA Code & Name", "Meaning", "Market",
           "Source View", "View Explanation", "Built?"]
ws.append(headers)
for c in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=c)
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
    cell.border = BORDER

for r in rows:
    ws.append(list(r))
    ridx = ws.max_row
    for c in range(1, 8):
        cell = ws.cell(row=ridx, column=c)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = BORDER
        cell.font = Font(size=10)
    ws.cell(row=ridx, column=1).fill = CAT_FILL
    ws.cell(row=ridx, column=1).font = Font(size=9, italic=True)
    ws.cell(row=ridx, column=2).font = Font(size=10, bold=True)
    ws.cell(row=ridx, column=7).fill = built_fill(r[6])

widths = [26, 34, 46, 8, 42, 50, 22]
for idx, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(idx)].width = w

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:G{ws.max_row}"

# ---- Sheet 2: Remarks Reconciliation (Gaurav's ACCOUNT_VIEW_MAP mark-up = DQ-33) ----
DONE = PatternFill("solid", fgColor="C6EFCE")
PARK = PatternFill("solid", fgColor="FFEB9C")
remarks = [
    ("1", "Verification charge is NOT revenue — it's cash we refund; strip off 4025, book like a deposit (2110 liability)",
     "DONE",
     "tmpl 68 charge → Dr 1019 / Cr 2110; tmpl 69 refund → Dr 2110 / Cr 1019 (was 5053). H1 AU nets to A$2.00 outstanding — never touches P&L. Mirrors deposit tmpls 50/51. (DQ-38)"),
    ("2", "Connect-clearing views are CASH not accrual — rename _a_ → _c_",
     "DONE",
     "Created view_{SG,AU}_c_host_rms_internal_clearing (byte-identical), repointed view_map (feeds both host_rms_internal_clearing + connect_internal_transfer), dropped old _a_ views. Ties unchanged (SG 48,798.92 / AU 13,605.11). (DQ-38)"),
    ("3", "RMS revenue split (4001) = by CONNECT ACCOUNTS, not host earnings",
     "DONE",
     "4001 fed by view_{SG,AU}_a_trip_revenue_rms — split by connect-account mechanism (ENT-7), same basis as the trip-revenue split. Map note corrected. (DQ-34)"),
    ("4", "AU fuel charge (4023) comes from INCIDENTAL INVOICES, not inside trip cash",
     "DONE",
     "4023 AU fed by view_AU_a_incidentals_fuel_charge_invoiced (classifier on incidental-invoice lines). (DQ-34)"),
    ("5", "4002 Flex+ = subscription + Flex+ incidentals invoices; 4003 Flex+ RMS = split",
     "DONE / PARKED",
     "4002 DONE — two feeds: subscription_flexplus_invoiced + _LEAK_FLEXPLUS incidentals carve-out (DQ-35). 4003 P2P/RMS split PARKED at your instruction (POL-61) — all Flex+ stays in 4002 for now; split must later cover BOTH feeds."),
    ("6", "4011 insurance subscription must be built / separated from total subscription",
     "DONE",
     "4011 fed by view_{SG,AU}_a_subscription_insurance_invoiced (+ insurance-leak carve-out); broad subscriptions_invoiced lump deactivated. (DQ-34)"),
    ("7", "5055 should NOT all fold into 5054 — split insurance-sub refunds out",
     "DONE",
     "Insurance-vs-device signal (lines LIKE '%insur%'). 5054 = device-only; NEW 5055 = view_{SG,AU}_c_subscription_refunds_insurance (tmpl 139/140). Ties: SG −4,087.88 (ins 0); AU device −6,548.30 + ins −4,736.17. (DQ-39)"),
    ("8", "Misc payouts are DESCRIPTION-ONLY, NEVER host-type / RMS split",
     "DONE",
     "misc_corrections_rms_clearing removed; AU trip/distance corrections all → 5000 single line; misc classifier is keyword-on-description only, never touches connect/mechanism logic. (DQ-33 #8, POL-56)"),
    ("9", "4025 = incidentals-invoice Other + direct guest payments; verification stripped",
     "DONE",
     "4025 fed by incidentals_other + direct_revenue + (new) subscription_other + AU paginated-hold; verification removed (remark #1). (DQ-33 #9, DQ-36)"),
]
ws2 = wb.create_sheet("Remarks Reconciliation")
h2 = ["#", "Your remark (from the marked-up file)", "Status", "What was done"]
ws2.append(h2)
for c in range(1, 5):
    cell = ws2.cell(row=1, column=c)
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = BORDER
for num, remark, status, did in remarks:
    ws2.append([num, remark, status, did])
    r = ws2.max_row
    for c in range(1, 5):
        cell = ws2.cell(row=r, column=c)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = BORDER
        cell.font = Font(size=10)
    ws2.cell(row=r, column=3).fill = PARK if "PARK" in status else DONE
    ws2.cell(row=r, column=3).font = Font(size=10, bold=True)
for idx, w in enumerate([5, 52, 16, 78], start=1):
    ws2.column_dimensions[get_column_letter(idx)].width = w
ws2.freeze_panes = "A2"

# put the reconciliation sheet first (it's the headline deliverable)
wb.move_sheet("Remarks Reconciliation", -(len(wb.sheetnames) - 1))

wb.save(OUT)

# quick summary
from collections import Counter
built = Counter()
for r in rows:
    s = r[6].lower()
    if s.startswith("built"):
        built["built"] += 1
    elif "external" in s:
        built["external"] += 1
    else:
        built["not-built"] += 1
print(f"rows={len(rows)}  built={built['built']}  external={built['external']}  not-built={built['not-built']}")
print(f"saved -> {OUT}")
