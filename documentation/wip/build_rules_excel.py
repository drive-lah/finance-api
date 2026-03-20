"""
Build a reviewable Excel sheet from QB rule CSV exports.
Output columns match the finance-api FinanceCategorizationRule schema.
"""
import csv
import os
from collections import defaultdict

BASE = os.path.dirname(__file__)

def read_csv(filename):
    path = os.path.join(BASE, filename)
    with open(path, newline='', encoding='utf-8') as f:
        return list(csv.DictReader(f))

# ── Load all 5 files ─────────────────────────────────────────────────────────
companies   = read_csv("qb_rules_companies_202603131707.csv")
rules       = read_csv("qb_rules_202603131707.csv")
conditions  = read_csv("qb_rule_conditions_202603131706.csv")
actions     = read_csv("qb_rule_actions_202603131706.csv")
# account mappings: not needed for the sheet — QB account path comes from action type 0 directly

# ── Lookup tables ─────────────────────────────────────────────────────────────
realm_to_entity = {c["realm_id"]: c["company_name"] for c in companies}

# Group conditions and actions by rule_id
conds_by_rule  = defaultdict(list)
for c in conditions:
    conds_by_rule[c["rule_id"]].append(c)

actions_by_rule = defaultdict(list)
for a in actions:
    actions_by_rule[a["rule_id"]].append(a)

# ── Bank account name mapping (QB name → readable + new ba ID hint) ──────────
# These are SOURCE bank account filters (which ba the rule is scoped to).
# Maps to bank_account_ids[] on the FinanceCategorizationRule.
BA_NAME_MAP = {
    "1001":       ("OCBC 1001", "1"),
    "3001":       ("OCBC 3001", "18"),
    "Wise":       ("Wise SGD",  "2"),
    "stripe_sgd": ("Stripe SGD","19"),
    "CBA":        ("CBA AU",    "? (AU CBA)"),
    "DBS_USD_Holding": ("DBS USD", "? (Ventures DBS USD)"),
    "DBS_SGD_Holding": ("DBS SGD", "? (Ventures DBS SGD)"),
    "accrual_sg": ("Accrual (non-bank)", "N/A"),
    "":           ("(all accounts)", ""),
}

def resolve_ba(raw):
    if not raw:
        return "", ""
    parts = [p.strip() for p in raw.split(",")]
    names = []
    ids   = []
    for p in parts:
        n, i = BA_NAME_MAP.get(p, (p, "?"))
        names.append(n)
        ids.append(i)
    return ", ".join(names), ", ".join(ids)

# ── Infer target_bank_account_id for INTERNAL_TRANSFER rules ─────────────────
# These are pre-filled from rule names where the destination is obvious.
# Maps to target_bank_account_id on the rule (destination bank account).
TRANSFER_TARGET_MAP = {
    # rule_name pattern keywords → target ba id
    "1001 incoming from 3001":  "1",    # 3001 → 1001, target=ba1
    "1001 outgoing to 3001":   "18",    # 1001 → 3001, target=ba18
    "1001 outgoing to wise":   "2",     # 1001 → Wise,  target=ba2
    "3001 incoming from 1001": "18",    # 1001 → 3001, target=ba18
    "3001 incoming from stripe":"18",   # Stripe → 3001, target=ba18
    "3001 outgoing to 1001":   "1",     # 3001 → 1001, target=ba1
    "3001 outgoing to wise":   "2",     # 3001 → Wise,  target=ba2
    "wise incoming from 1001": "2",     # 1001 → Wise,  target=ba2
    "wise incoming from 3001": "2",     # 3001 → Wise,  target=ba2
    "stripe outgoing to 3001": "18",    # Stripe → 3001, target=ba18
    "dbs usd outgoing to dbs sgd": "? (Ventures DBS SGD)",
    "dbs sgd incoming from dbs usd": "? (Ventures DBS SGD)",
}

def get_transfer_target(rule_name, category):
    if category != "INTERNAL_TRANSFER":
        return ""
    lower = rule_name.lower()
    for pattern, ba_id in TRANSFER_TARGET_MAP.items():
        if pattern in lower:
            return ba_id
    return "?"   # INTERNAL_TRANSFER but target not auto-detected — human fills

# ── Derive direction from condition type 10 ───────────────────────────────────
def get_direction(rule_id):
    for c in conds_by_rule.get(rule_id, []):
        if c["condition_type"] == "10":
            v = c["condition_value"].strip()
            if v == "-1":
                return "OUTGOING"
            elif v == "1":
                return "INCOMING"
    return ""

# ── Derive description values from condition types 1 and 6 ───────────────────
def get_desc_values(rule_id):
    vals = []
    for c in conds_by_rule.get(rule_id, []):
        if c["condition_type"] in ("1", "6"):
            v = c["condition_value"].strip()
            if v and v not in vals:
                vals.append(v)
    return " | ".join(vals)

# ── Get QB account name from action type 0 ────────────────────────────────────
def get_qb_account(rule_id):
    for a in actions_by_rule.get(rule_id, []):
        if a["action_type"] == "0":
            return a["action_value"].strip()
    return ""

# ── Get counterparty name from action type 5 ─────────────────────────────────
def get_counterparty(rule_id):
    for a in actions_by_rule.get(rule_id, []):
        if a["action_type"] == "5":
            return a["action_value"].strip()
    return ""

# ── Infer category hint ───────────────────────────────────────────────────────
def infer_category(direction, qb_account, rule_name):
    name_lower  = rule_name.lower()
    acct_lower  = qb_account.lower()
    if "transfer" in name_lower:
        return "INTERNAL_TRANSFER"
    if direction == "INCOMING":
        return "DEPOSIT"
    if direction == "OUTGOING":
        return "EXPENSE"
    if "income" in acct_lower or "revenue" in acct_lower:
        return "DEPOSIT"
    if "expense" in acct_lower or "cost" in acct_lower or "fee" in acct_lower:
        return "EXPENSE"
    return ""


# ── COA lookup: QB account path keyword → (code, name, confidence) ───────────
# Confidence: HIGH = very confident, MEDIUM = likely correct, LOW = needs review
QB_ACCOUNT_KEYWORD_MAP = [
    # Technology
    ("technology system cost",          "6701", "Technology - Software Subscriptions", "HIGH"),
    ("technology system costs",         "6701", "Technology - Software Subscriptions", "HIGH"),
    ("tech operations",                 "6700", "Technology - Infrastructure",         "HIGH"),
    # Marketing
    ("marketing - advertising",         "6100", "Marketing - Digital Advertising",     "HIGH"),
    ("marketing - asset creation",      "6103", "Marketing - Asset Creation",          "HIGH"),
    ("marketing - discount",            "5050", "Monthly Discounts",                   "MEDIUM"),
    # Office / Corporate
    ("corporate - office rent",         "6300", "Office Rent",                         "HIGH"),
    ("corporate - bank fees",           "6600", "Bank Fees",                           "HIGH"),
    ("corporate - bookkeeping fees",    "6500", "Accounting & Bookkeeping Fees",       "HIGH"),
    ("bookkeeping fees",                "6500", "Accounting & Bookkeeping Fees",       "HIGH"),
    ("accounting and bookkeeping",      "6500", "Accounting & Bookkeeping Fees",       "HIGH"),
    ("legal and professional fees",     "6501", "Legal Fees",                          "HIGH"),
    ("corporate - legal",               "6501", "Legal Fees",                          "HIGH"),
    ("corporate-other corporate",       "6014", "Employee Claims - Other",             "LOW"),
    ("other corporate expense",         "6014", "Employee Claims - Other",             "LOW"),
    # Salary / HR
    ("salary",                          "6000", "Salaries & Wages",                    "HIGH"),
    ("employees expense",               "6000", "Salaries & Wages",                    "HIGH"),
    ("human resource",                  "6200", "HR - Recruitment",                    "MEDIUM"),
    # Travel
    ("travel expense",                  "6400", "Travel - Tickets",                    "HIGH"),
    ("meals and entertainment",         "6402", "Entertainment",                       "HIGH"),
    # Fleet / Host payouts
    ("due from fleet",                  "1220", "Other Receivables",                   "MEDIUM"),
    ("host earnings",                   "5000", "Host Payouts - P2P",                  "HIGH"),
    ("host payouts",                    "5000", "Host Payouts - P2P",                  "HIGH"),
    ("payment gateway",                 "5010", "Payment Processing Fees",             "HIGH"),
    ("stripe monthly",                  "5010", "Payment Processing Fees",             "HIGH"),
    # Revenue
    ("gross revenue-car rental",        "4000", "GBV - P2P",                           "HIGH"),
    ("gross revenue",                   "4000", "GBV - P2P",                           "MEDIUM"),
    ("billable expense income",         "4025", "Incidentals Revenue - Other",         "MEDIUM"),
    ("device subscriptions",            "4010", "Subscription Revenue - Device",       "HIGH"),
    ("flex+",                           "4002", "GBV - Flex+",                         "HIGH"),
    # Stripe deposits
    ("stripe - customer deposits",      "2110", "Customer Deposits Held",              "HIGH"),
    ("stripe - verification deposits",  "2110", "Customer Deposits Held",              "HIGH"),
    ("customer deposits",               "2110", "Customer Deposits Held",              "HIGH"),
    # Refunds
    ("refunds",                         "5052", "Refunds - Trip",                      "MEDIUM"),
    # Incidentals / fleet costs
    ("sticker payouts",                 "5041", "Host Payouts - Sticker",              "HIGH"),
    ("towing",                          "5033", "Incidentals Payout - Towing",         "HIGH"),
    ("late return",                     "5044", "Host Payouts - Late Return",          "HIGH"),
    ("misc dispute",                    "5025", "Incidentals Payout - Other",          "MEDIUM"),
    ("parking",                         "5060", "Parking - RMS Fleet",                 "HIGH"),
    ("fuel",                            "5023", "Incidentals Payout - Fuel",           "HIGH"),
    # Insurance
    ("cost of insurance",               "5031", "Cost of Insurance - Subscription Premium", "MEDIUM"),
    ("insurance",                       "5031", "Cost of Insurance - Subscription Premium", "LOW"),
    # IC
    ("due from australian",             "8000", "IC - Due from AU (SG books)",         "HIGH"),
    ("due from australia",              "8000", "IC - Due from AU (SG books)",         "HIGH"),
    ("due from singapore",              "8010", "IC - Due from SG (AU books)",         "HIGH"),
    ("due to australia",                "8100", "IC - Due to AU (SG books)",           "HIGH"),
    ("due to singapore",                "8110", "IC - Due to SG (AU books)",           "HIGH"),
    ("due to ventures",                 "8101", "IC - Due to Ventures (SG books)",     "HIGH"),
    ("due from ventures",               "8001", "IC - Due from Ventures (SG books)",   "HIGH"),
    ("amount owing to subsidiary",      "8100", "IC - Due to AU (SG books)",           "LOW"),
    ("amount owing by holding",         "8001", "IC - Due from Ventures (SG books)",   "LOW"),
    # Depreciation / Amortisation
    ("amortisation",                    "7400", "Amortisation - Technology Development","LOW"),
    ("amortization",                    "7400", "Amortisation - Technology Development","LOW"),
    ("depreciation",                    "7300", "Depreciation - Computer & Equipment", "LOW"),
    ("accumulated depreciation",        "1600", "Accum Depr - Computer & Peripherals", "LOW"),
    # Tax
    ("tax paid",                        "9000", "Income Tax Expense",                  "HIGH"),
    ("income tax",                      "9000", "Income Tax Expense",                  "HIGH"),
    ("ato",                             "9000", "Income Tax Expense",                  "MEDIUM"),
    # Misc
    ("exchange gain or loss",           "7100", "Foreign Exchange Gains/Losses",       "HIGH"),
    ("currency exchange",               "1330", "Currency Exchange",                   "HIGH"),
    ("ccy conversion",                  "6600", "Bank Fees",                           "HIGH"),
    ("reconciliation discrepancies",    "3300", "Suspense Account",                    "HIGH"),
    ("suspense",                        "3300", "Suspense Account",                    "HIGH"),
    ("cash rebate",                     "7001", "Other Income - Cash Rebate",          "HIGH"),
    ("loans to others",                 "1320", "Loans to Others",                     "HIGH"),
    ("prepaid",                         "1300", "Prepayments",                         "HIGH"),
    ("deferred income",                 "2210", "Deferred Income",                     "HIGH"),
    ("other income",                    "4025", "Incidentals Revenue - Other",         "LOW"),
    ("cost of overhead",                "5062", "On-Ground Team - Expenses",           "LOW"),
    ("cost of service",                 "5000", "Host Payouts - P2P",                  "LOW"),
    ("accrued expense",                 "2200", "Accrued Expenses",                    "HIGH"),
    ("accruals",                        "2200", "Accrued Expenses",                    "HIGH"),
    ("accounts payable",                "2000", "Trade & Other Payables",              "HIGH"),
    ("wise new",                        "",     "(internal transfer target)",           "LOW"),
]

# Rule-name keyword overrides — checked before QB account path
RULE_NAME_KEYWORD_MAP = [
    # SaaS infrastructure
    ("aws",         "6700", "Technology - Infrastructure",         "HIGH"),
    ("amazon web",  "6700", "Technology - Infrastructure",         "HIGH"),
    ("google cloud","6700", "Technology - Infrastructure",         "HIGH"),
    ("heroku",      "6700", "Technology - Infrastructure",         "HIGH"),
    ("digital matter","6700","Technology - Infrastructure",        "HIGH"),
    ("kore wireless","6700","Technology - Infrastructure",         "HIGH"),
    ("gps malaysia","6700", "Technology - Infrastructure",         "HIGH"),
    ("liberty wire","6700", "Technology - Infrastructure",         "HIGH"),
    # SaaS subscriptions
    ("slack",       "6701", "Technology - Software Subscriptions", "HIGH"),
    ("notion",      "6701", "Technology - Software Subscriptions", "HIGH"),
    ("canva",       "6701", "Technology - Software Subscriptions", "HIGH"),
    ("adobe",       "6701", "Technology - Software Subscriptions", "HIGH"),
    ("github",      "6701", "Technology - Software Subscriptions", "HIGH"),
    ("circleci",    "6701", "Technology - Software Subscriptions", "HIGH"),
    ("intercom",    "6701", "Technology - Software Subscriptions", "HIGH"),
    ("zendesk",     "6701", "Technology - Software Subscriptions", "HIGH"),
    ("twilio",      "6701", "Technology - Software Subscriptions", "HIGH"),
    ("hotjar",      "6701", "Technology - Software Subscriptions", "HIGH"),
    ("mixpanel",    "6701", "Technology - Software Subscriptions", "HIGH"),
    ("metabase",    "6701", "Technology - Software Subscriptions", "HIGH"),
    ("moengage",    "6701", "Technology - Software Subscriptions", "HIGH"),
    ("hubspot",     "6701", "Technology - Software Subscriptions", "HIGH"),
    ("atlassian",   "6701", "Technology - Software Subscriptions", "HIGH"),
    ("zapier",      "6701", "Technology - Software Subscriptions", "HIGH"),
    ("docusign",    "6701", "Technology - Software Subscriptions", "HIGH"),
    ("hellosign",   "6701", "Technology - Software Subscriptions", "HIGH"),
    ("typeform",    "6701", "Technology - Software Subscriptions", "HIGH"),
    ("lucidchart",  "6701", "Technology - Software Subscriptions", "HIGH"),
    ("papertrail",  "6701", "Technology - Software Subscriptions", "HIGH"),
    ("surveymonkey","6701", "Technology - Software Subscriptions", "HIGH"),
    ("calendly",    "6701", "Technology - Software Subscriptions", "HIGH"),
    ("dropbox",     "6701", "Technology - Software Subscriptions", "HIGH"),
    ("loom",        "6701", "Technology - Software Subscriptions", "HIGH"),
    ("miro",        "6701", "Technology - Software Subscriptions", "HIGH"),
    ("monday",      "6701", "Technology - Software Subscriptions", "HIGH"),
    ("bitly",       "6701", "Technology - Software Subscriptions", "HIGH"),
    ("overflow",    "6701", "Technology - Software Subscriptions", "HIGH"),
    ("onelink",     "6701", "Technology - Software Subscriptions", "HIGH"),
    ("murf.ai",     "6701", "Technology - Software Subscriptions", "HIGH"),
    ("midjourney",  "6701", "Technology - Software Subscriptions", "HIGH"),
    ("openai",      "6701", "Technology - Software Subscriptions", "HIGH"),
    ("chatgpt",     "6701", "Technology - Software Subscriptions", "HIGH"),
    ("paddle",      "6701", "Technology - Software Subscriptions", "HIGH"),
    ("retool",      "6701", "Technology - Software Subscriptions", "HIGH"),
    ("upwork",      "6701", "Technology - Software Subscriptions", "MEDIUM"),
    ("voucherify",  "6701", "Technology - Software Subscriptions", "HIGH"),
    ("sentry",      "6701", "Technology - Software Subscriptions", "HIGH"),
    ("uptime",      "6701", "Technology - Software Subscriptions", "HIGH"),
    ("aircall",     "6701", "Technology - Software Subscriptions", "HIGH"),
    ("gupshup",     "6701", "Technology - Software Subscriptions", "HIGH"),
    ("onfido",      "6701", "Technology - Software Subscriptions", "HIGH"),
    ("recruitee",   "6701", "Technology - Software Subscriptions", "HIGH"),
    ("illion",      "6701", "Technology - Software Subscriptions", "HIGH"),
    ("adjust",      "6701", "Technology - Software Subscriptions", "HIGH"),
    ("sharetribe",  "6701", "Technology - Software Subscriptions", "HIGH"),
    ("godaddy",     "6701", "Technology - Software Subscriptions", "HIGH"),
    ("xero",        "6500", "Accounting & Bookkeeping Fees",       "HIGH"),
    ("intuit",      "6500", "Accounting & Bookkeeping Fees",       "HIGH"),
    ("octopods",    "6701", "Technology - Software Subscriptions", "HIGH"),
    ("timetastic",  "6701", "Technology - Software Subscriptions", "HIGH"),
    ("birthdaybot", "6701", "Technology - Software Subscriptions", "HIGH"),
    ("notionforms", "6701", "Technology - Software Subscriptions", "HIGH"),
    ("docsend",     "6701", "Technology - Software Subscriptions", "HIGH"),
    ("trust pilot", "6701", "Technology - Software Subscriptions", "HIGH"),
    ("finger print","6701", "Technology - Software Subscriptions", "HIGH"),
    ("safetyculture","6701","Technology - Software Subscriptions", "HIGH"),
    ("apple.com",   "6701", "Technology - Software Subscriptions", "HIGH"),
    ("microsoft",   "6701", "Technology - Software Subscriptions", "HIGH"),
    ("google suite","6701", "Technology - Software Subscriptions", "HIGH"),
    ("google play", "6701", "Technology - Software Subscriptions", "HIGH"),
    ("rightworks",  "6701", "Technology - Software Subscriptions", "HIGH"),
    ("focussed claims","6501","Legal Fees",                        "HIGH"),
    ("apex debt",   "6501", "Legal Fees",                         "HIGH"),
    ("legal fees",  "6501", "Legal Fees",                         "HIGH"),
    # Marketing
    ("facebook",    "6100", "Marketing - Digital Advertising",    "HIGH"),
    ("google ads",  "6100", "Marketing - Digital Advertising",    "HIGH"),
    ("marketing new customers", "6100","Marketing - Digital Advertising","HIGH"),
    ("carousell",   "6100", "Marketing - Digital Advertising",    "MEDIUM"),
    ("photo",       "6103", "Marketing - Asset Creation",         "MEDIUM"),
    # Office
    ("justco",      "6300", "Office Rent",                        "HIGH"),
    ("bank charges","6600", "Bank Fees",                          "HIGH"),
    ("wise transfer fee","6600","Bank Fees",                      "HIGH"),
    # HR / Salary
    ("contractor salary","6000","Salaries & Wages",               "HIGH"),
    ("wise corporate salary","6000","Salaries & Wages",           "HIGH"),
    ("wise hr salary","6000","Salaries & Wages",                  "HIGH"),
    ("cpf",         "2300", "CPF Payable (SG)",                   "HIGH"),
    ("ntuc",        "6011", "Employee Claims - Meals",            "MEDIUM"),
    ("income insurance","6000","Salaries & Wages",               "LOW"),
    ("ph healthcare","6014","Employee Claims - Other",            "MEDIUM"),
    ("caretaker",   "6000", "Salaries & Wages",                   "MEDIUM"),
    ("human resource","6200","HR - Recruitment",                  "MEDIUM"),
    # Fleet
    ("rms",         "5001", "Host Payouts - P2P RMS",             "MEDIUM"),
    ("eurokars",    "5001", "Host Payouts - P2P RMS",             "HIGH"),
    ("alpine rms",  "5001", "Host Payouts - P2P RMS",             "HIGH"),
    ("komoco",      "5001", "Host Payouts - P2P RMS",             "HIGH"),
    ("cdg rms",     "5001", "Host Payouts - P2P RMS",             "HIGH"),
    ("cycle & carria","5001","Host Payouts - P2P RMS",            "HIGH"),
    ("cheng auto",  "5062", "On-Ground Team - Expenses",          "MEDIUM"),
    ("arena detailing","5062","On-Ground Team - Expenses",        "MEDIUM"),
    ("battery company","5062","On-Ground Team - Expenses",        "MEDIUM"),
    ("city auto",   "5062", "On-Ground Team - Expenses",          "MEDIUM"),
    ("drive mate",  "5001", "Host Payouts - P2P RMS",             "MEDIUM"),
    # Transfers / IC
    ("money received from australia","8000","IC - Due from AU (SG books)","HIGH"),
    ("australia transfer","8000","IC - Due from AU (SG books)",   "MEDIUM"),
    # Tax (AU)
    ("ato",         "9000", "Income Tax Expense",                 "HIGH"),
    ("tax payments","9000", "Income Tax Expense",                 "HIGH"),
    # Other
    ("parking.sg",  "5060", "Parking - RMS Fleet",               "HIGH"),
    ("misc earnings","4025","Incidentals Revenue - Other",        "LOW"),
    ("general debit","6014","Employee Claims - Other",            "LOW"),
    ("refunds on card","7001","Other Income - Cash Rebate",       "LOW"),
    ("cash rebate", "7001", "Other Income - Cash Rebate",         "HIGH"),
    ("accounting expense","6500","Accounting & Bookkeeping Fees", "HIGH"),
    ("aussie broadband","6300","Office Rent",                     "MEDIUM"),
    ("starbucks",   "6401", "Travel - Meals",                     "HIGH"),
    ("kaolin",      "5062", "On-Ground Team - Expenses",          "LOW"),
    ("ibg giro",    "6000", "Salaries & Wages",                   "LOW"),
    ("transfer agent","6014","Employee Claims - Other",           "LOW"),
    # Stripe cash rules (SG)
    ("c_stripe_fees_paid",          "5010", "Payment Processing Fees",             "HIGH"),
    ("c_disputes",                  "5051", "Chargebacks",                         "HIGH"),
    ("c_stripe_payouts",            "",     "(bank-to-bank payout — skip)",         "LOW"),
    ("c_trip_cash_collected",       "2100", "Deferred Trip Revenue",               "HIGH"),
    ("c_incidentals_invoice_paid",  "1200", "Trade Receivables",                   "HIGH"),
    ("c_subscription_invoice_paid", "1200", "Trade Receivables",                   "HIGH"),
    ("c_customer_deposits",         "2110", "Customer Deposits Held",              "HIGH"),
    ("c_host_transfers",            "",     "(host payout — INTERNAL_TRANSFER)",    "LOW"),
    # SG Accrual rules (non-bank, revenue recognition — likely out of scope)
    ("a_trip_revenue_earned",       "4000", "GBV - P2P",                           "MEDIUM"),
    ("a_subscription_invoiced",     "4010", "Subscription Revenue - Device",       "MEDIUM"),
    ("a_incidentals_invoiced",      "4025", "Incidentals Revenue - Other",         "MEDIUM"),
    ("a_host_trip_earnings",        "5000", "Host Payouts - P2P",                  "MEDIUM"),
    ("a_host_incidentals_damage",   "5021", "Incidentals Payout - Damage (Host)",  "MEDIUM"),
    ("a_host_incidentals_excess",   "5024", "Incidentals Payout - Excess Mileage", "MEDIUM"),
    ("a_host_incidentals_tolls",    "5020", "Incidentals Payout - Tolls",          "MEDIUM"),
    ("a_host_incidentals_clean",    "5022", "Incidentals Payout - Cleaning",       "MEDIUM"),
    ("a_host_misc_payout",          "5042", "Host Payouts - Misc",                 "MEDIUM"),
    ("a_host_flexplus_payout",      "5002", "Host Payouts - Flex+",                "MEDIUM"),
    ("a_host_superhost_payout",     "5040", "Host Payouts - Superhost",            "MEDIUM"),
    ("a_host_sticker_payout",       "5041", "Host Payouts - Sticker",              "MEDIUM"),
    ("a_host_referral_payout",      "5042", "Host Payouts - Misc",                 "LOW"),
    ("a_host_subscription",         "5030", "Cost of Device Subscriptions",        "LOW"),
    ("a_host_misc_charge",          "4025", "Incidentals Revenue - Other",         "LOW"),
    ("a_trip_distance",             "4000", "GBV - P2P",                           "MEDIUM"),
    # AU Stripe cash rules
    ("c_stripe_payouts",            "",     "(bank-to-bank payout — skip)",         "LOW"),
    # Other specific
    ("journey horizon",             "1710", "Technology Development",              "LOW"),
    ("fleet earnings",              "2120", "Host Payables",                       "MEDIUM"),
    ("due to fleet",                "2120", "Host Payables",                       "MEDIUM"),
    ("bank fees international",     "6600", "Bank Fees",                           "HIGH"),
    ("international transaction",   "6600", "Bank Fees",                           "HIGH"),
    ("linkt",                       "5020", "Incidentals Payout - Tolls",          "MEDIUM"),
    ("marketing asset creation",    "6103", "Marketing - Asset Creation",          "HIGH"),
    ("chargefox",                   "5060", "Parking - RMS Fleet",                 "MEDIUM"),
    ("mazda",                       "5001", "Host Payouts - P2P RMS",              "MEDIUM"),
    ("wise new",                    "",     "(internal transfer — skip)",           "LOW"),
    ("c_trip_distance_cash",        "4000", "GBV - P2P",                           "MEDIUM"),
    ("c_trip_distance_invoice",     "1200", "Trade Receivables",                   "MEDIUM"),
]

def derive_contra(qb_account, rule_name, desc_values, category):
    """Return (code, name, confidence) for the contra account."""
    if category == "INTERNAL_TRANSFER":
        return ("", "", "N/A - INTERNAL_TRANSFER")

    combined = (rule_name + " " + qb_account + " " + desc_values).lower()
    rule_lower = rule_name.lower()

    # 1. Rule-name keyword overrides (most specific)
    for kw, code, name, conf in RULE_NAME_KEYWORD_MAP:
        if kw in rule_lower:
            return (code, name, conf)

    # 2. QB account path keyword matching
    qb_lower = qb_account.lower()
    for kw, code, name, conf in QB_ACCOUNT_KEYWORD_MAP:
        if kw in qb_lower:
            return (code, name, conf)

    # 3. Fallback: check combined text
    for kw, code, name, conf in QB_ACCOUNT_KEYWORD_MAP:
        if kw in combined:
            return (code, name, conf)

    return ("", "", "MANUAL")

# ── Build rows ────────────────────────────────────────────────────────────────
rows = []
for r in rules:
    rule_id    = r["id"]
    realm      = r["realm_id"]
    rule_name  = r["rule_name"]
    desc_note  = r.get("description", "")
    priority   = r["priority"]
    is_active  = r["is_active"]
    ba_raw     = r.get("bank_accounts", "").strip()

    entity_name       = realm_to_entity.get(realm, realm)
    ba_name, ba_id    = resolve_ba(ba_raw)
    direction         = get_direction(rule_id)
    desc_values       = get_desc_values(rule_id)
    qb_account        = get_qb_account(rule_id)
    counterparty_name = get_counterparty(rule_id)
    category_hint     = infer_category(direction, qb_account, rule_name)

    target_ba = get_transfer_target(rule_name, category_hint)
    contra_code, contra_name, confidence = derive_contra(
        qb_account, rule_name, desc_values, category_hint
    )

    # Flag accrual rules — these are non-bank accounting entries, likely out of scope
    extra_note = ""
    if "accrual" in rule_name.lower() and not desc_values:
        extra_note = "⚠️ Accrual/non-bank rule — may not apply to bank recon engine"
    elif "c_stripe_payouts" in rule_name.lower():
        extra_note = "⚠️ Stripe payout to bank — likely INTERNAL_TRANSFER, not expense"

    rows.append({
        "qb_rule_id":               rule_id,
        "entity":                   entity_name,
        "rule_name":                rule_name,
        # SOURCE bank account — which ba(s) this rule is scoped to (bank_account_ids[])
        "source_bank_account_name": ba_name,
        "source_bank_account_id":   ba_id,
        "priority":                 priority,
        "is_active":                is_active,
        "direction":                direction,
        "description_values":       desc_values,
        "description_operator":     "CONTAINS" if desc_values else "",
        "category":                 category_hint,
        "qb_account_name":          qb_account,
        # Derived contra account — human to confirm/correct
        "contra_account_code":      contra_code,
        "contra_account_name":      contra_name,
        "confidence":               confidence,
        # TARGET bank account — only for INTERNAL_TRANSFER (target_bank_account_id)
        "target_bank_account_id":   target_ba,
        "counterparty_name":        counterparty_name,
        "notes":                    (extra_note + " " + desc_note).strip(),
    })

# ── Write Excel ───────────────────────────────────────────────────────────────
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Rules"

HEADERS = [
    "qb_rule_id", "entity", "rule_name",
    # SOURCE — which bank account(s) the rule is scoped to
    "source_bank_account_name", "source_bank_account_id",
    "priority", "is_active",
    "direction",
    "description_values", "description_operator",
    "category",
    "qb_account_name",           # QB path for reference
    # Derived contra account — confirm/correct
    "contra_account_code",
    "contra_account_name",
    "confidence",                # HIGH/MEDIUM/LOW/MANUAL/N/A
    # TARGET — only for INTERNAL_TRANSFER: destination bank account
    "target_bank_account_id",
    "counterparty_name",
    "notes",
]

# Columns user reviews / corrects (yellow)
FILL_COLS = {"contra_account_code", "contra_account_name", "category", "target_bank_account_id"}

HEADER_FILL   = PatternFill("solid", fgColor="1F497D")
HEADER_FONT   = Font(bold=True, color="FFFFFF")
EDIT_FILL     = PatternFill("solid", fgColor="FFF2CC")   # yellow  = human to confirm
REF_FILL      = PatternFill("solid", fgColor="DEEAF1")   # blue    = read-only ref
HIGH_FILL     = PatternFill("solid", fgColor="E2EFDA")   # green   = HIGH confidence
MEDIUM_FILL   = PatternFill("solid", fgColor="FFEB9C")   # amber   = MEDIUM confidence
LOW_FILL      = PatternFill("solid", fgColor="FFC7CE")   # red     = LOW / MANUAL needs review
NA_FILL       = PatternFill("solid", fgColor="D9D9D9")   # grey    = N/A (transfers)

CONF_COL_IDX  = HEADERS.index("confidence") + 1

ws.append(HEADERS)
for col_idx, h in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.fill  = HEADER_FILL
    cell.font  = HEADER_FONT
    cell.alignment = Alignment(horizontal="center")

for row_data in rows:
    ws.append([row_data[h] for h in HEADERS])

# Colour the data rows
for row_idx in range(2, len(rows) + 2):
    conf = ws.cell(row_idx, CONF_COL_IDX).value or ""
    if conf == "HIGH":
        contra_fill = HIGH_FILL
    elif conf == "MEDIUM":
        contra_fill = MEDIUM_FILL
    elif conf in ("LOW", "MANUAL", ""):
        contra_fill = LOW_FILL
    else:
        contra_fill = NA_FILL   # N/A - INTERNAL_TRANSFER

    for col_idx, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        if h in ("contra_account_code", "contra_account_name", "confidence"):
            cell.fill = contra_fill
        elif h in FILL_COLS:
            cell.fill = EDIT_FILL
        else:
            cell.fill = REF_FILL

# Auto-size columns
for col_idx, h in enumerate(HEADERS, 1):
    col_letter = get_column_letter(col_idx)
    max_len = len(h)
    for row_idx in range(2, len(rows) + 2):
        v = ws.cell(row=row_idx, column=col_idx).value or ""
        max_len = max(max_len, len(str(v)))
    ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

# Freeze header row
ws.freeze_panes = "A2"

# Add a "COA Reference" sheet so user can see current COA codes
ws2 = wb.create_sheet("COA Reference")
ws2.append(["account_code", "account_name", "account_type", "category"])
ws2.cell(row=1, column=1).font = Font(bold=True)
ws2.cell(row=1, column=2).font = Font(bold=True)
ws2.cell(row=1, column=3).font = Font(bold=True)
ws2.cell(row=1, column=4).font = Font(bold=True)
ws2.append(["-- Pull from GET /api/finance/accounts --", "", "", ""])

# Save
out_path = os.path.join(BASE, "rules_for_review_v4.xlsx")
wb.save(out_path)
print(f"Written {len(rows)} rules → {out_path}")
