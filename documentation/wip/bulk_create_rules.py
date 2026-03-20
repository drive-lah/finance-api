#!/usr/bin/env python3
"""
Bulk create categorization rules from Excel sheet.
Reads rules_for_review_v4.xlsx, creates rules via API, tracks results.
"""
import sys
import os
import json
import requests
from pathlib import Path
from decimal import Decimal
from datetime import datetime

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Load environment
from dotenv import load_dotenv
load_dotenv()

API_BASE = "http://localhost:8081/api/finance"
EXCEL_PATH = "/Users/gauravsinghal/Documents/Work/G-master/finance-api/documentation/wip/rules_for_review_v4.xlsx"
COMPANIES_CSV = "/Users/gauravsinghal/Documents/Work/G-master/finance-api/documentation/wip/qb_rules_companies_202603131707.csv"

# Column mappings to Excel (actual names)
COLUMNS = {
    "qb_rule_id": "qb_rule_id",
    "entity_name": "entity",
    "rule_name": "rule_name",
    "source_bank_account_name": "source_bank_account_name",
    "source_bank_account_id": "source_bank_account_id",
    "priority": "priority",
    "is_active": "is_active",
    "direction": "direction",
    "description_value": "description_values",
    "description_operator": "description_operator",
    "category": "category",
    "qb_account_name": "qb_account_name",
    "contra_account_code": "contra_account_code",
    "target_bank_account_id": "target_bank_account_id",
    "counterparty_name": "counterparty_name",
}

def load_company_mapping():
    """Load company_name → entity_id mapping from CSV."""
    companies_df = pd.read_csv(COMPANIES_CSV)
    mapping = dict(zip(companies_df["company_name"], companies_df["id"]))
    print(f"Loaded {len(mapping)} companies: {list(mapping.keys())}")
    return mapping

def read_excel_rules():
    """Read rules from Excel, return rows that are filled."""
    df = pd.read_excel(EXCEL_PATH, sheet_name=0)
    print(f"Total rows in Excel: {len(df)}")

    # Filter rows where contra_account_code or target_bank_account_id is filled
    # (indicating the rule has been configured by user)
    filled = df[
        (df[COLUMNS["contra_account_code"]].notna()) |
        (df[COLUMNS["target_bank_account_id"]].notna())
    ].reset_index(drop=True)

    print(f"Filled rules (with contra or target_bank): {len(filled)}")
    return filled

def build_rule_payload(row, company_mapping):
    """Convert Excel row to RuleCreate JSON payload."""
    entity_name = row[COLUMNS["entity_name"]]
    entity_id = company_mapping.get(entity_name)

    if not entity_id:
        raise ValueError(f"Unknown entity: {entity_name}")

    direction = str(row[COLUMNS["direction"]]).upper().strip() if pd.notna(row[COLUMNS["direction"]]) else "INCOMING"
    category = str(row[COLUMNS["category"]]).upper().strip() if pd.notna(row[COLUMNS["category"]]) else "EXPENSE"

    description_value = str(row[COLUMNS["description_value"]]).strip() if pd.notna(row[COLUMNS["description_value"]]) else None
    description_operator = str(row[COLUMNS["description_operator"]]).upper() if pd.notna(row[COLUMNS["description_operator"]]) else None
    if description_value and not description_operator:
        description_operator = "CONTAINS"

    counterparty_value = str(row[COLUMNS["counterparty_name"]]).strip() if pd.notna(row[COLUMNS["counterparty_name"]]) else None
    counterparty_operator = "CONTAINS" if counterparty_value else None

    target_bank_id = None
    if pd.notna(row[COLUMNS["target_bank_account_id"]]):
        try:
            target_bank_id = int(float(row[COLUMNS["target_bank_account_id"]]))
        except (ValueError, TypeError):
            pass

    contra_code = None
    if pd.notna(row[COLUMNS["contra_account_code"]]):
        contra_code = str(row[COLUMNS["contra_account_code"]]).strip()
        if contra_code == "nan":
            contra_code = None

    payload = {
        "name": f"{entity_name} - {str(row[COLUMNS['rule_name']]).strip()}",
        "priority": int(row[COLUMNS["priority"]]) if pd.notna(row[COLUMNS["priority"]]) else 100,
        "status": "ACTIVE" if row[COLUMNS["is_active"]] else "INACTIVE",
        "description": f"QB Rule: {str(row[COLUMNS['qb_rule_id']])} - {str(row[COLUMNS['qb_account_name']]).strip() if pd.notna(row[COLUMNS['qb_account_name']]) else 'N/A'}",

        # Scope
        "direction": direction,
        "bank_account_ids": None,

        # Match criteria
        "amount_value": None,
        "description_value": description_value,
        "description_operator": description_operator,
        "counterparty_value": counterparty_value,
        "counterparty_operator": counterparty_operator,
        "counterparty_id": None,
        "match_currency": None,

        # Action
        "category": category,
        "contra_account_code": contra_code,
        "target_bank_account_id": target_bank_id,
        "allocation_entity_id": None,
        "tag_ids": None,
        "gst_override": None,
    }

    # Clean up None values for optional fields
    payload = {k: v for k, v in payload.items() if v is not None}

    return payload

def create_rule(payload):
    """POST rule to API, return response or error."""
    try:
        resp = requests.post(f"{API_BASE}/rules", json=payload, timeout=10)
        resp.raise_for_status()
        return resp.json()
    except requests.exceptions.RequestException as e:
        return {"error": str(e), "status_code": resp.status_code if 'resp' in locals() else None}

def bulk_create(filled_df, company_mapping):
    """Bulk create rules and track results."""
    results = []
    created_count = 0
    error_count = 0

    for idx, row in filled_df.iterrows():
        rule_name = row[COLUMNS["rule_name"]]
        try:
            payload = build_rule_payload(row, company_mapping)
            result = create_rule(payload)

            if "error" in result:
                print(f"❌ Row {idx}: {rule_name} — {result['error']}")
                results.append({
                    "index": idx,
                    "rule_name": rule_name,
                    "status": "ERROR",
                    "rule_id": None,
                    "error": result.get("error"),
                })
                error_count += 1
            else:
                rule_id = result.get("id")
                print(f"✅ Row {idx}: {rule_name} → rule_id={rule_id}")
                results.append({
                    "index": idx,
                    "rule_name": rule_name,
                    "status": "CREATED",
                    "rule_id": rule_id,
                    "error": None,
                })
                created_count += 1
        except Exception as e:
            print(f"❌ Row {idx}: {rule_name} — Exception: {str(e)}")
            results.append({
                "index": idx,
                "rule_name": rule_name,
                "status": "ERROR",
                "rule_id": None,
                "error": str(e),
            })
            error_count += 1

    return results, created_count, error_count

def update_excel(results):
    """Update Excel with rule_id and status columns."""
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    created_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
    error_fill = PatternFill(start_color="FF6347", end_color="FF6347", fill_type="solid")    # Red

    for result in results:
        row_num = result["index"] + 2  # +2 for header and 0-indexing
        rule_id = result["rule_id"]
        status = result["status"]

        # Column P (16) = rule_id, Column O (15) = status
        if rule_id:
            ws.cell(row=row_num, column=16).value = rule_id
        ws.cell(row=row_num, column=15).value = status

        # Color the status cell
        fill = created_fill if status == "CREATED" else error_fill
        ws.cell(row=row_num, column=15).fill = fill

    wb.save(EXCEL_PATH)
    print(f"\n✅ Excel updated: {EXCEL_PATH}")

def main():
    print("=" * 80)
    print("BULK CREATE CATEGORIZATION RULES")
    print("=" * 80)

    # Load company mapping
    company_mapping = load_company_mapping()

    # Read filled rules from Excel
    filled_df = read_excel_rules()
    if len(filled_df) == 0:
        print("❌ No filled rules found in Excel")
        return

    # Bulk create rules
    print(f"\nCreating {len(filled_df)} rules...")
    results, created, errors = bulk_create(filled_df, company_mapping)

    # Update Excel with results
    update_excel(results)

    # Summary
    print("\n" + "=" * 80)
    print(f"SUMMARY: {created} created, {errors} errors")
    print("=" * 80)

    if errors > 0:
        print("\nFailed rules:")
        for r in results:
            if r["status"] == "ERROR":
                print(f"  - {r['rule_name']}: {r['error']}")

if __name__ == "__main__":
    main()
