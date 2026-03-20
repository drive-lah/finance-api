#!/usr/bin/env python3
"""
Bulk create categorization rules directly via SQLAlchemy.
Reads rules_for_review_v4.xlsx, creates rules in DB, tracks results in Excel.
"""
import sys
import os
from pathlib import Path
from datetime import datetime

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from dotenv import load_dotenv
load_dotenv()

# Import models
from src.models.categorization_rule import FinanceCategorizationRule, RuleStatus, TransactionDirection, TransactionCategory, MatchOperator

EXCEL_PATH = "/Users/gauravsinghal/Documents/Work/G-master/finance-api/documentation/wip/rules_for_review_v4.xlsx"
COMPANIES_CSV = "/Users/gauravsinghal/Documents/Work/G-master/finance-api/documentation/wip/qb_rules_companies_202603131707.csv"
DATABASE_URL = os.getenv('DATABASE_URL', 'postgresql://localhost/finance_db')

# Column mappings to Excel (actual names)
COLUMNS = {
    "qb_rule_id": "qb_rule_id",
    "entity_name": "entity",
    "rule_name": "rule_name",
    "source_bank_account_name": "source_bank_account_name",
    "source_bank_account_id": "source_bank_account_id",
    "priority": "priority",
    "is_active": "is_active",
    "direction": "direction",
    "description_value": "description_values",
    "description_operator": "description_operator",
    "category": "category",
    "qb_account_name": "qb_account_name",
    "contra_account_code": "contra_account_code",
    "target_bank_account_id": "target_bank_account_id",
    "counterparty_name": "counterparty_name",
}

def load_company_mapping():
    """Load company_name → entity_id mapping from CSV."""
    companies_df = pd.read_csv(COMPANIES_CSV)
    mapping = dict(zip(companies_df["company_name"], companies_df["id"]))
    print(f"Loaded {len(mapping)} companies")
    return mapping

def read_excel_rules():
    """Read rules from Excel, return rows that are filled."""
    df = pd.read_excel(EXCEL_PATH, sheet_name=0)
    print(f"Total rows in Excel: {len(df)}")

    # Filter rows where contra_account_code or target_bank_account_id is filled
    filled = df[
        (df[COLUMNS["contra_account_code"]].notna()) |
        (df[COLUMNS["target_bank_account_id"]].notna())
    ].reset_index(drop=True)

    print(f"Filled rules (with contra or target_bank): {len(filled)}")
    return filled

def build_rule_obj(row, company_mapping):
    """Convert Excel row to FinanceCategorizationRule object."""
    entity_name = row[COLUMNS["entity_name"]]
    if entity_name not in company_mapping:
        raise ValueError(f"Unknown entity: {entity_name}")

    direction_str = str(row[COLUMNS["direction"]]).upper().strip() if pd.notna(row[COLUMNS["direction"]]) else "INCOMING"
    direction = TransactionDirection[direction_str]

    category_str = str(row[COLUMNS["category"]]).upper().strip() if pd.notna(row[COLUMNS["category"]]) else "EXPENSE"
    category = TransactionCategory[category_str]

    description_value = str(row[COLUMNS["description_value"]]).strip() if pd.notna(row[COLUMNS["description_value"]]) else None
    description_operator_str = str(row[COLUMNS["description_operator"]]).upper() if pd.notna(row[COLUMNS["description_operator"]]) else None
    if description_value and not description_operator_str:
        description_operator_str = "CONTAINS"
    description_operator = MatchOperator[description_operator_str] if description_operator_str else None

    counterparty_value = str(row[COLUMNS["counterparty_name"]]).strip() if pd.notna(row[COLUMNS["counterparty_name"]]) else None
    counterparty_operator = MatchOperator.CONTAINS if counterparty_value else None

    target_bank_id = None
    if pd.notna(row[COLUMNS["target_bank_account_id"]]):
        try:
            target_bank_id = int(float(row[COLUMNS["target_bank_account_id"]]))
        except (ValueError, TypeError):
            pass

    contra_code = None
    if pd.notna(row[COLUMNS["contra_account_code"]]):
        contra_code = str(row[COLUMNS["contra_account_code"]]).strip()
        if contra_code.lower() not in ("nan", "none", ""):
            pass
        else:
            contra_code = None

    rule = FinanceCategorizationRule(
        name=f"{entity_name} - {str(row[COLUMNS['rule_name']]).strip()}",
        priority=int(row[COLUMNS["priority"]]) if pd.notna(row[COLUMNS["priority"]]) else 100,
        status=RuleStatus.ACTIVE if row[COLUMNS["is_active"]] else RuleStatus.INACTIVE,
        description=f"QB Rule: {str(row[COLUMNS['qb_rule_id']])} - {str(row[COLUMNS['qb_account_name']]).strip() if pd.notna(row[COLUMNS['qb_account_name']]) else 'N/A'}",

        direction=direction,
        bank_account_ids=None,

        amount_value=None,
        amount_operator=None,
        amount_value_max=None,
        description_value=description_value,
        description_operator=description_operator,
        transaction_type_value=None,
        transaction_type_operator=None,
        counterparty_id=None,
        counterparty_value=counterparty_value,
        counterparty_operator=counterparty_operator,
        counterparty_name=counterparty_value,
        counterparty_type=None,
        match_currency=None,

        category=category,
        contra_account_code=contra_code,
        target_bank_account_id=target_bank_id,
        allocation_entity_id=None,
        tag_ids=None,
        gst_override=None,
    )

    return rule

def bulk_create_db(filled_df, company_mapping):
    """Create rules in database and return results."""
    engine = create_engine(DATABASE_URL)
    results = []
    created_count = 0
    error_count = 0

    with Session(engine) as session:
        for idx, row in filled_df.iterrows():
            rule_name = row[COLUMNS["rule_name"]]
            try:
                rule_obj = build_rule_obj(row, company_mapping)
                session.add(rule_obj)
                session.commit()

                rule_id = rule_obj.id
                print(f"✅ Row {idx}: {rule_name} → rule_id={rule_id}")
                results.append({
                    "index": idx,
                    "rule_name": rule_name,
                    "status": "CREATED",
                    "rule_id": rule_id,
                    "error": None,
                })
                created_count += 1
            except Exception as e:
                session.rollback()
                error_msg = str(e)
                print(f"❌ Row {idx}: {rule_name} — {error_msg}")
                results.append({
                    "index": idx,
                    "rule_name": rule_name,
                    "status": "ERROR",
                    "rule_id": None,
                    "error": error_msg,
                })
                error_count += 1

    return results, created_count, error_count

def update_excel(results):
    """Update Excel with rule_id and status columns."""
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    created_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    error_fill = PatternFill(start_color="FF6347", end_color="FF6347", fill_type="solid")

    for result in results:
        row_num = result["index"] + 2  # +2 for header and 0-indexing
        rule_id = result["rule_id"]
        status = result["status"]

        # Append columns if they don't exist yet
        # Column S (19) = status, Column T (20) = rule_id
        ws.cell(row=row_num, column=19).value = status
        if rule_id:
            ws.cell(row=row_num, column=20).value = rule_id

        # Color the status cell
        fill = created_fill if status == "CREATED" else error_fill
        ws.cell(row=row_num, column=19).fill = fill

    wb.save(EXCEL_PATH)
    print(f"\n✅ Excel updated: {EXCEL_PATH}")

def main():
    print("=" * 80)
    print("BULK CREATE CATEGORIZATION RULES (Direct DB)")
    print("=" * 80)

    # Load company mapping
    company_mapping = load_company_mapping()

    # Read filled rules from Excel
    filled_df = read_excel_rules()
    if len(filled_df) == 0:
        print("❌ No filled rules found in Excel")
        return

    # Bulk create rules
    print(f"\nCreating {len(filled_df)} rules in database...")
    results, created, errors = bulk_create_db(filled_df, company_mapping)

    # Update Excel with results
    update_excel(results)

    # Summary
    print("\n" + "=" * 80)
    print(f"SUMMARY: {created} created, {errors} errors")
    print("=" * 80)

    if errors > 0:
        print("\nFailed rules:")
        for r in results:
            if r["status"] == "ERROR":
                print(f"  - {r['rule_name']}: {r['error']}")

if __name__ == "__main__":
    main()
