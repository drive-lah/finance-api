#!/usr/bin/env python3
"""Rename map for the naming-consistency pass: every USED view -> its v_ canonical name.

Core transform (safe, mechanical): view_ -> v_ ; drop trailing _new.
Also emits an OPTIONAL deeper-normalization suggestion per view (not applied here).
Nothing is renamed — this only produces the review map.
"""
import re, json, subprocess, glob, csv
from collections import defaultdict
from src.services.economic_events.view_map import VIEW_MAP, PAYOUT_LINE_VIEWS
from src.database import get_session_factory
from sqlalchemy import text
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CH = 'http://54.169.212.254:8123/?user=clickhouse-server-drivelah&password=Drivelah2025'
def ch(q): return subprocess.run(['curl', '-s', CH, '--data-binary', q], capture_output=True, text=True).stdout
views = {o['name']: o['create_table_query'] for o in (json.loads(r) for r in ch(
    "SELECT name,create_table_query FROM system.tables WHERE database='default' AND engine='View' AND name LIKE 'view_%' FORMAT JSONEachRow").strip().split('\n') if r.strip())}
allviews = set(views)

s = get_session_factory()()
ent = {'SG': 2, 'AU': 3}
active_ets = {(eid, et) for et, eid in s.execute(text("SELECT event_type,entity_id FROM finance_je_templates WHERE is_active"))}
s.close()
active_ee = {sp.view for (m, et), sp in VIEW_MAP.items() if (ent.get(m), et) in active_ets}
payout_lane = {sp.view for sp in PAYOUT_LINE_VIEWS.values()}

deps = defaultdict(set)
for v, sql in views.items():
    for o in allviews:
        if o != v and re.search(r'\b' + re.escape(o) + r'\b', sql):
            deps[v].add(o)
def clo(seeds):
    seen = set(); st = list(seeds)
    while st:
        for f in deps.get(st.pop(), ()):
            if f not in seen:
                seen.add(f); st.append(f)
    return seen
roots = active_ee | payout_lane
feeders = clo(roots) - roots
used = roots | feeders

def used_as(v):
    if v in active_ee: return "economic-events"
    if v in payout_lane: return "payout lane"
    return "feeder"

def core_new(v):
    n = "v_" + v[len("view_"):]      # view_ -> v_
    if n.endswith("_new"): n = n[:-4]  # drop _new
    return n

def further(v, newname):
    tips = []
    if "_internal_clearing" in v or ("_c_" in v and "clearing" in v):
        tips.append("lane -> _x_ (internal/clearing, no P&L)")
    for q in ("_invoiced", "_paid", "_collected", "_received"):
        if newname.endswith(q):
            tips.append(f"drop redundant '{q}' (lane letter already says it)")
    if "host_payout" in v or "host_trip_earnings" in v or "host_superhost" in v:
        tips.append("unify host-cost verb (host_<category>)")
    if "host_incidentals_" in v:
        tips.append("host_incidentals_X -> host_X (parallel with revenue)")
    return " ; ".join(tips)

rows = []
seen_new = defaultdict(list)
for v in sorted(used):
    n = core_new(v)
    seen_new[n].append(v)
    change = "prefix+drop_new" if v.endswith("_new") else "prefix"
    rows.append([v, n, change, used_as(v), further(v, n)])

# collision check
collisions = {n: vs for n, vs in seen_new.items() if len(vs) > 1}

with open('documentation/wip/VIEW_RENAME_MAP.csv', 'w', newline='') as f:
    w = csv.writer(f); w.writerow(["Old view", "New (v_ canonical)", "Change", "Used as", "Optional further normalization"])
    for r in rows: w.writerow(r)

wb = Workbook(); ws = wb.active; ws.title = "Rename Map"
HF = PatternFill("solid", fgColor="1F4E78"); TH = Side(style="thin", color="BFBFBF"); BD = Border(TH, TH, TH, TH)
hdr = ["Old view", "New (v_ canonical)", "Change", "Used as", "Optional further normalization"]
ws.append(hdr)
for c in range(1, 6):
    cell = ws.cell(1, c); cell.font = Font(bold=True, color="FFFFFF"); cell.fill = HF
    cell.alignment = Alignment(wrap_text=True, vertical="center"); cell.border = BD
fill = {"prefix": "C6EFCE", "prefix+drop_new": "FFF2CC"}
for r in rows:
    ws.append(r); rr = ws.max_row
    for c in range(1, 6):
        cell = ws.cell(rr, c); cell.border = BD; cell.alignment = Alignment(wrap_text=True, vertical="top"); cell.font = Font(size=10)
    ws.cell(rr, 3).fill = PatternFill("solid", fgColor=fill.get(r[2], "FFFFFF"))
for i, w in enumerate([48, 44, 16, 18, 46], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:E{ws.max_row}"
wb.save('documentation/wip/VIEW_RENAME_MAP.xlsx')

print(f"used views: {len(used)} (economic-events {len(active_ee)}, payout {len(payout_lane)}, feeders {len(feeders)})")
print(f"  prefix-only: {sum(1 for r in rows if r[2]=='prefix')}  |  prefix+drop_new: {sum(1 for r in rows if r[2]=='prefix+drop_new')}")
print(f"COLLISIONS (new names claimed by >1 old view): {collisions if collisions else 'NONE'}")
print("saved VIEW_RENAME_MAP.csv + .xlsx")
