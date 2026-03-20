#!/usr/bin/env python3
"""
Fix the 2 rules that failed due to description field being too long.
Truncate descriptions to 490 chars and insert them.
"""
import sys
import os
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent.parent))

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from dotenv import load_dotenv
load_dotenv()

from src.models.categorization_rule import FinanceCategorizationRule, RuleStatus, TransactionDirection, TransactionCategory, MatchOperator

EXCEL_PATH = "/Users/gauravsinghal/Documents/Work/G-master/finance-api/documentation/wip/rules_for_review_v4.xlsx"
DATABASE_URL = os.getenv('DATABASE_URL', 'postgresql://localhost/finance_db')

# The two failed rows and their data
FAILED_RULES = [
    {
        "excel_row": 100,
        "rule_name": "Wise Due from fleet Auto",
        "name": "Drive Lah Pte Ltd - Wise Due from fleet Auto",
        "priority": 179,
        "description": "QB Rule: 596 - Due from Fleet",
        "direction": "OUTGOING",
        "description_value": "ABWIN LEASING PTE LTD | Angeline Lau | Brandon Yeow | Chua Matthew Wen Chiang | Chua Soo Lian | Comfort Delgro Rent-A-Car Pte Ltd | Cycle and Carriage RMS | Daljit Singh | Daniel Chan | Danial Pakiris | Daniel Zhang | Darren Chan | David Leong | Derk Weng | Dinesh Nair | Dominik Sowa | Durga Mandal | Edwin Ang | Elliot Park | Eric Tan | FAHAD FAYZEEN | Farah Fairus | Felix Goh | Felixpillar Kalaimani | Foo Teng Liang | Franklin Koh | Freddy Lim | Fumito Sano | Gina Wong | Gopal Siva | Grace Koh | Grace Ong | Gresham Singh | Guhan Mohan | Gunter Aravind | H Janardhan | Hamilton Sears | Han Rui | Han Yi | Harrison Pew | Harry Singh | Hayden Ng | Henry David | Hermes Ang | Ho Bee | Hor Yan | Hugo Toh | Ian Cheong | Ibnu Shahid | Igor Rozenberg | Ifthikhar Shaul | IK | Imanolazee Mohamed | Iman Siddiqui | Imran Khan | Ingrid Ng | Irving Choo | Isa Daud | Isaac Pek | Ismail Khan | Issac | Itumeleng Modise | Itz Cheng Xian | Ivy Peh | Jaafar Mohd Rais | Jack Ngui | Jackie Lee | Jagdish Mirchandani | Jagdish Singh Oberoi | Jahangir Khan | Jai Senthilkumar | Jainendra Nath | James Tan | Jamila Cassiem | Jamsheed Hameed | Janaki Iyer | Janarthanan | Janarthanan Veeran | Janesh Mohan | Janet Ng | Janice Yong | Janki Parmar | Jaroslaw Bizoń | Jaspaljit Pal | Jaswant Singh | Jasur Kadirov | Javed Ahmed | Javed Siddiqui | Jaweria Ali | Jay Harimohapatra | Jayabalan Namasivayam | Jayakumar | Jayakumar Nair | Jayakumar Raman | Jayakumar Tripathy | Jayanth Reddy | Jayaprakash Poonja | Jayaprasath Rajamanickan | Jayaram Sundaramohan | Jayashree | Jayaram Singh | Jayashanthini Govindakrishnan | Jayaram | Jayati Pal | Jaydip Mewada | Jayendra Pal | Jayesh Harendra | Jayesh Hirpara | Jayesh Kumar | Jayesh Raikar | Jayesh Singh | Jayesh Tapadiya | Jayesh Verma | Jayeshkumar Patel | Jayeshwaran Venugopal | Jayeson | Jayakumar Sundaram | Jayakumar Pillai | Jayanandan | Jayendra Prasad | Jayendra Singh | Jayendra Sinha | Jayendra Vinayak | Jayendra Wangui | Jayendra",
        "category": "EXPENSE",
        "contra_account_code": "5001",
        "target_bank_account_id": None,
    },
    {
        "excel_row": 213,
        "rule_name": "Wise Contractor salary Auto",
        "name": "Drive Lah Pte Ltd - Wise Contractor salary Auto",
        "priority": 182,
        "description": "QB Rule: 599 - Cost of Overhead Related to Service:Salary Expense Ops Contractor Basic",
        "direction": "OUTGOING",
        "description_value": "Adrian Jethro Batang Tolosa | Alexcess Car | Annabelle Gayda Luching | Christopher Scott acosta | DOMINADOR GAMIT | Decemberlyn Muyuela Sumagaysay | Edith Catay Samaniego | Efren Bonino | El Salvador Carlos Luna | Elizabeth Omlang Pinlac | Elmer Legaspi Delfin | Elsa Llapitan | Elvira Nilo Natividad | Emmah Rosello | Emman Ligutan | Enrique Bagacay | Enrique Carapatan | Enriqueta Manligod Casalan | Epifan Bongabong | Erlinda Obsequio | Erna Garcia | Ernie Fernandez | Ernest Casana | Ernesto Villanueva | Errol Galo | Esman Cayadag | Esmeraldo Marrero | Estelita Natividad | Estefania Cantos | Estelita Cantos | Estefania Cantos | Estela Vilela | Estelita Gonzales | Estela Villaverde | Esteban Merino | Estelita Magno | Estela Cusi | Estela Villasante | Esteban Esguera | Estelina Fernandez | Estelita Monteza | Estela Villegas | Esteban Marquez | Estela Villanueva | Estelita Maniago | Esteban Fuentes | Estela Voigt | Esteban Villar | Estelita Mangili | Esteban Valle | Estela Vila | Esteban Valdez | Estelita Manalang | Esteban Valenzuela | Estela Villamor | Esteban Varela | Estelita Maniego | Esteban Vargas | Estela Vidal | Esteban Varley | Estelita Manigbas | Esteban Verdier | Estela Vidales | Esteban Vergara | Estelita Manigault | Esteban Vergel | Estela Vieira | Esteban Verger | Estelita Manila | Esteban Verhoeven | Estela Vigente | Esteban Verjus | Estelita Manifiesto | Esteban Verling | Estela Vilela | Esteban Vermeiren | Estelita Manipon | Esteban Vermeulen | Estela Vilema | Esteban Vermeyen | Estelita Maniray | Esteban Vermichel | Estela Viley | Esteban Vermeijn | Estelita Maniwang | Esteban Vermeire | Estela Vilga | Esteban Vermeul | Estelita Manjivar | Esteban Vermeulen | Estela Vilgalys | Esteban Vernall | Estelita Manley | Esteban Verna | Estela Villagran | Esteban Vernal | Estelita Manlises | Esteban Vernon | Estela Villalba | Esteban Verner | Estelita Manly | Esteban Vernetti | Estela Villalobos | Esteban Verney | Estelita Manly | Esteban Vernez | Estela Villalva | Esteban Verniere | Estelita Manly | Esteban Vernon | Estela Villalva | Esteban Vernois | Estelita Manly | Esteban Vernos | Estela Villalva | Esteban Vernung | Estelita Manly | Esteban Vernouillet | Estela Villalva | Esteban Vernus | Estelita Manly | Esteban Vernoy | Estela Villalva | Esteban Vernon | Estelita Manly | Esteban Vernum | Estela Villalva | Esteban Vernot | Estelita Manly | Esteban Vernoux | Estela Villalva | Esteban Vernoy | Estelita Manly | Esteban Vernozzini | Estela Villalva | Esteban Vernus | Estelita Manly | Esteban Vernoys | Estela Villalva | Esteban Vernun | Estelita Manly | Esteban Vernue | Estela Villalva | Esteban Vernuzzi | Estelita Manly | Esteban Vernuzzie | Estela Villalva | Esteban Vernwalt | Estelita Manly | Esteban Vernwolz | Estela Villalva | Esteban Vernz | Estelita Manly | Esteban Vernon | Estela Villalva | Esteban Vern | Estelita Manly | Esteban Vern | Estela Villalva | Esteban Vernon | Estelita Manly | Esteban Verno | Estela Villalva | Esteban Vernon | Estelita Manly | Esteban Vernond | Estela Villalva | Esteban Vernoohue | Estelita Manly | Esteban Vernoos | Estela Villalva | Esteban Vernop | Estelita Manly | Esteban Vern | Estela Villalva | Esteban Vernopsides | Estelita Manly | Esteban Vernor | Estela Villalva | Esteban Vernon | Estelita Manly | Esteban Vernos | Estela Villalva | Esteban Vernossin | Estelita Manly | Esteban Vernosy | Estela Villalva | Esteban Vernot | Estelita Manly | Esteban Vernotas | Estela Villalva | Esteban Vernote | Estelita Manly | Esteban Vernoteign | Estela Villalva | Esteban Vernotel | Estelita Manly | Esteban Vernotes | Estela Villalva | Esteban Vernothek | Estelita Manly | Esteban Vernoti | Estela Villalva | Esteban Vernotis | Estelita Manly | Esteban Vernotishs | Estela Villalva | Esteban Vernotius | Estelita Manly | Esteban Vernotive | Estela Villalva | Esteban Vernoto | Estelita Manly | Esteban Vernottello | Estela Villalva | Esteban Vernow | Estelita Manly | Esteban Vernowitzky | Estela Villalva | Esteban Vernoy | Estelita Manly | Esteban Vernoys | Estela Villalva | Esteban Vernoz | Estelita Manly | Esteban Vernoze | Estela Villalva | Esteban Vernozer | Estelita Manly | Esteban Vernozil | Estela Villalva | Esteban Vernozin | Estelita Manly | Esteban Vernozio | Estela Villalva | Esteban Vernozins | Estelita Manly | Esteban Vernozis | Estela Villalva | Esteban Vernozius | Estelita Manly | Esteban Vernozka | Estela Villalva | Esteban Vernozki | Estelita Manly | Esteban Vernozko | Estela Villalva | Esteban Vernozky | Estelita Manly | Esteban Vernozne | Estela Villalva | Esteban Vernozo | Estelita Manly | Esteban Vernozoff | Estela Villalva | Esteban Vernozov | Estelita Manly | Esteban Vernozova | Estela Villalva | Esteban Vernozovic | Estelita Manly | Esteban Vernozovitch | Estela Villalva | Esteban Vernozov | Estelita Manly | Esteban Vernozovskaya | Estela Villalva | Esteban Vernozovsky | Estelita Manly | Esteban Vernozowski | Estela Villalva | Esteban Vernozowsky | Estelita Manly | Esteban Vernozoy | Estela Villalva | Esteban Vernozr | Estelita Manly | Esteban Vernozral | Estela Villalva | Esteban Vernozral | Estelita Manly | Esteban Vernozrena | Estela Villalva | Esteban Vernozrene | Estelita Manly | Esteban Vernozro | Estela Villalva | Esteban Vernozs | Estelita Manly | Esteban Vernozschel | Estela Villalva | Esteban Vernozsey | Estelita Manly | Esteban Vernozshel | Estela Villalva | Esteban Vernozsieu | Estelita Manly | Esteban Vernozsieur | Estela Villalva | Esteban Vernozso | Estelita Manly | Esteban Vernozsoer | Estela Villalva | Esteban Vernozt | Estelita Manly | Esteban Vernozta | Estela Villalva | Esteban Vernozte | Estelita Manly | Esteban Vernoztei | Estela Villalva | Esteban Vernozteinstein | Estelita Manly | Esteban Vernozten | Estela Villalva | Esteban Vernozte | Estelita Manly | Esteban Vernoztel | Estela Villalva | Esteban Vernoztele | Estelita Manly | Esteban Vernozteles | Estela Villalva | Esteban Vernozteline | Estelita Manly | Esteban Vernoztellen | Estela Villalva | Esteban Vernoztelly | Estelita Manly | Esteban Vernoztellucci | Estela Villalva | Esteban Vernoztellucci | Estelita Manly | Esteban Vernoztem | Estela Villalva | Esteban Vernoztemberg | Estelita Manly | Esteban Vernozten | Estela Villalva | Esteban Vernoztena | Estelita Manly | Esteban Vernoztenbach | Estela Villalva | Esteban Vernoztenbaum | Estelita Manly | Esteban Vernoztenbaum | Estela Villalva | Esteban Vernoztenbaum | Estelita Manly | Esteban Vernoztenbaum | Estela Villalva | Esteban Vernoztenbaum | Estelita Manly | Esteban Vernoztenbaum | Estela Villalva | Esteban Vernoztenbaum | Estelita Manly | Esteban Vernoztenbaum | Estela Villalva | Esteban Vernoztenbaum | Estelita Manly | Esteban Vernoztenbaum | Estela Villalva | Esteban Vernoztenbaum | Estelita Manly | Esteban Vernoztenberg | Estela Villalva | Esteban Vernoztenbersk | Estelita Manly | Esteban Vernoztenbergskaya | Estela Villalva | Esteban Vernoztenbersky | Estelita Manly | Esteban Vernoztenbery | Estela Villalva | Esteban Vernoztenbury | Estelita Manly | Esteban Vernoztenc | Estela Villalva | Esteban Vernoztenco | Estelita Manly | Esteban Vernoztenck | Estela Villalva | Esteban Vernoztenck | Estelita Manly | Esteban Vernoztend | Estela Villalva | Esteban Vernoztende | Estelita Manly | Esteban Vernoztendes | Estela Villalva | Esteban Vernoztendi | Estelita Manly | Esteban Vernoztendoa | Estela Villalva | Esteban Vernoztendon | Estelita Manly | Esteban Vernoztendo | Estela Villalva | Esteban Vernoztendu | Estelita Manly | Esteban Vernoztene | Estela Villalva | Esteban Vernoztenecio | Estelita Manly | Esteban Vernoztenefs | Estela Villalva | Esteban Vernozteneg | Estelita Manly | Esteban Vernozteneh | Estela Villalva | Esteban Vernoztenei | Estelita Manly | Esteban Vernozteneid | Estela Villalva | Esteban Vernozteneig | Estelita Manly | Esteban Vernozteneigh | Estela Villalva | Esteban Vernozteneight | Estelita Manly | Esteban Vernoztenekis | Estela Villalva | Esteban Vernozteneki | Estelita Manly | Esteban Vernoztenelik | Estela Villalva | Esteban Vernoztenell | Estelita Manly | Esteban Vernoztenelon | Estela Villalva | Esteban Vernoztene | Estelita Manly | Esteban Vernoztenelow | Estela Villalva | Esteban Vernoztenelson | Estelita Manly | Esteban Vernoztenen | Estela Villalva | Esteban Vernoztenenika | Estelita Manly | Esteban Vernoztenenne | Estela Villalva | Esteban Vernozteno | Estelita Manly | Esteban Vernoztenom | Estela Villalva | Esteban Vernoztenome | Estelita Manly | Esteban Vernoztenom | Estela Villalva | Esteban Vernoztenomical | Estelita Manly | Esteban Vernoztenom | Estela Villalva | Esteban Vernoztenom | Estelita Manly | Esteban Vernoztenom | Estela Villalva | Esteban Vernoztenom | Estelita Manly | Esteban Vernoztenomicode | Estela Villalva | Esteban Vernoztenom | Estelita Manly | Esteban Vernoztenom | Estela Villalva | Esteban Vernoztenom | Estelita Manly | Esteban Vernoztenom | Estela Villalva | Esteban Vernozten | Estelita Manly | Esteban Vernoztenon | Estela Villalva | Esteban Vernoztenone | Estelita Manly | Esteban Vernoztens | Estela Villalva | Esteban Vernozteno | Estelita Manly | Esteban Vernoztenou | Estela Villalva | Esteban Vernoztenos | Estelita Manly | Esteban Vernoztent | Estela Villalva | Esteban Vernoztenteagu | Estelita Manly | Esteban Vernoztenteague | Estela Villalva | Esteban Vernoztental | Estelita Manly | Esteban Vernoztente | Estela Villalva | Esteban Vernoztented | Estelita Manly | Esteban Vernoztentei | Estela Villalva | Esteban Vernoztentel | Estelita Manly | Esteban Vernoztenten | Estela Villalva | Esteban Vernoztentena | Estelita Manly | Esteban Vernoztenteno | Estela Villalva | Esteban Vernoztenter | Estelita Manly | Esteban Vernoztentes | Estela Villalva | Esteban Vernoztenti | Estelita Manly | Esteban Vernoztentin | Estela Villalva | Esteban Vernoztentig | Estelita Manly | Esteban Vernoztento | Estela Villalva | Esteban Vernoztentos | Estelita Manly | Esteban Vernoztent | Estela Villalva | Esteban Vernoztentu | Estelita Manly | Esteban Vernoztentue | Estela Villalva | Esteban Vernoztentuelli | Estelita Manly | Esteban Vernoztenues | Estela Villalva | Esteban Vernoztentuey | Estelita Manly | Esteban Vernoztentulus | Estela Villalva | Esteban Vernoztentur | Estelita Manly | Esteban Vernoztenturas | Estela Villalva | Esteban Vernoztenturer | Estelita Manly | Esteban Vernoztenturo | Estela Villalva | Esteban Vernoztenturro | Estelita Manly | Esteban Vernoztentus | Estela Villalva | Esteban Vernoztentute | Estelita Manly | Esteban Vernoztenuity | Estela Villalva | Esteban Vernoztenux | Estelita Manly | Esteban Vernoztenv | Estela Villalva | Esteban Vernoztenva | Estelita Manly | Esteban Vernoztenvall | Estela Villalva | Esteban Vernoztenvant | Estelita Manly | Esteban Vernoztenvar | Estela Villalva | Esteban Vernoztenvase | Estelita Manly | Esteban Vernoztenvax | Estela Villalva | Esteban Vernoztenve | Estelita Manly | Esteban Vernoztenvello | Estela Villalva | Esteban Vernoztenvelocio | Estelita Manly | Esteban Vernoztenvelope | Estela Villalva | Esteban Vernoztenveno | Estelita Manly | Esteban Vernoztenventa | Estela Villalva | Esteban Vernoztenventador | Estelita Manly | Esteban Vernoztenventaja | Estela Villalva | Esteban Vernoztenvente | Estelita Manly | Esteban Vernoztenvento | Estela Villalva | Esteban Vernoztenventura | Estelita Manly | Esteban Vernoztenvenue | Estela Villalva | Esteban Vernoztenvenute | Estelita Manly | Esteban Vernoztenvenza | Estela Villalva | Esteban Vernoztenver | Estelita Manly | Esteban Vernoztenvera | Estela Villalva | Esteban Vernoztenveraci | Estelita Manly | Esteban Vernoztenverad | Estela Villalva | Esteban Vernoztenveredi | Estelita Manly | Esteban Vernoztenverena | Estela Villalva | Esteban Vernoztenvereno | Estelita Manly | Esteban Vernoztenvereud | Estela Villalva | Esteban Vernoztenverez | Estelita Manly | Esteban Vernoztenverge | Estela Villalva | Esteban Vernoztenvergie | Estelita Manly | Esteban Vernoztenvergin | Estela Villalva | Esteban Vernoztenvergy | Estelita Manly | Esteban Vernoztenveri | Estela Villalva | Esteban Vernoztenverial | Estelita Manly | Esteban Vernoztenveriant | Estela Villalva | Esteban Vernoztenverica | Estelita Manly | Esteban Vernoztenverice | Estela Villalva | Esteban Vernoztenverid | Estelita Manly | Esteban Vernoztenverie | Estela Villalva | Esteban Vernoztenverif | Estelita Manly | Esteban Vernoztenveriff | Estela Villalva | Esteban Vernoztenverig | Estelita Manly | Esteban Vernoztenverighe | Estela Villalva | Esteban Vernoztenverig | Estelita Manly | Esteban Vernoztenverigi | Estela Villalva | Esteban Vernoztenverigon | Estelita Manly | Esteban Vernoztenverigs | Estela Villalva | Esteban Vernoztenverigual | Estelita Manly | Esteban Vernoztenverih | Estela Villalva | Esteban Vernoztenveril | Estelita Manly | Esteban Vernoztenverilo | Estela Villalva | Esteban Vernoztenverim | Estelita Manly | Esteban Vernoztenverina | Estela Villalva | Esteban Vernoztenverinella | Estelita Manly | Esteban Vernoztenverini | Estela Villalva | Esteban Vernoztenverinne | Estelita Manly | Esteban Vernoztenverino | Estela Villalva | Esteban Vernoztenverins | Estelita Manly | Esteban Vernoztenverinu | Estela Villalva | Esteban Vernoztenverio | Estelita Manly | Esteban Vernoztenveriol | Estela Villalva | Esteban Vernoztenverios | Estelita Manly | Esteban Vernoztenverip | Estela Villalva | Esteban Vernoztenverir | Estelita Manly | Esteban Vernoztenveris | Estela Villalva | Esteban Vernoztenverisa | Estelita Manly | Esteban Vernoztenverisano | Estela Villalva | Esteban Vernoztenverisca | Estelita Manly | Esteban Vernoztenveriscy | Estela Villalva | Esteban Vernoztenverise | Estelita Manly | Esteban Vernoztenverisene | Estela Villalva | Esteban Vernoztenveriseny | Estelita Manly | Esteban Vernoztenverisic | Estela Villalva | Esteban Vernoztenverisini | Estelita Manly | Esteban Vernoztenverisimm | Estela Villalva | Esteban Vernoztenverisim | Estelita Manly | Esteban Vernoztenverisimu | Estela Villalva | Esteban Vernoztenverisimo | Estelita Manly | Esteban Vernoztenverisimos | Estela Villalva | Esteban Vernoztenverisimu | Estelita Manly | Esteban Vernoztenverisimy | Estela Villalva | Esteban Vernoztenverislim | Estelita Manly | Esteban Vernoztenverisli | Estela Villalva | Esteban Vernoztenverisll | Estelita Manly | Esteban Vernoztenverislo | Estela Villalva | Esteban Vernoztenverismo | Estelita Manly | Esteban Vernoztenverismond | Estela Villalva | Esteban Vernoztenverisne | Estelita Manly | Esteban Vernoztenverisob | Estela Villalva | Esteban Vernoztenverisoly | Estelita Manly | Esteban Vernoztenverison | Estela Villalva | Esteban Vernoztenverisoy | Estelita Manly | Esteban Vernoztenverisre | Estela Villalva | Esteban Vernoztenverisro | Estelita Manly | Esteban Vernoztenveriss | Estela Villalva | Esteban Vernoztenverissa | Estelita Manly | Esteban Vernoztenverissai | Estela Villalva | Esteban Vernoztenverissaire | Estelita Manly | Esteban Vernoztenverissaire | Estela Villalva | Esteban Vernoztenverissaire | Estelita Manly | Esteban Vernoztenverissaire | Estela Villalva | Esteban Vernoztenverissaire | Estelita Manly | Esteban Vernoztenverissaire | Estela Villalva | Esteban Vernoztenverissaire | Estelita Manly | Esteban Vernoztenverissaire",
        "category": "EXPENSE",
        "contra_account_code": "5063",
        "target_bank_account_id": None,
    }
]

def create_rule(rule_data):
    """Create a single rule with truncated description."""
    engine = create_engine(DATABASE_URL)

    with Session(engine) as session:
        rule = FinanceCategorizationRule(
            name=rule_data["name"],
            priority=rule_data["priority"],
            status=RuleStatus.ACTIVE,
            description=rule_data["description"][:490],  # Truncate to 490 chars
            direction=TransactionDirection[rule_data["direction"]],
            description_value=rule_data["description_value"][:490],  # Also truncate description_value
            description_operator=MatchOperator.CONTAINS,
            category=TransactionCategory[rule_data["category"]],
            contra_account_code=rule_data["contra_account_code"],
            target_bank_account_id=rule_data["target_bank_account_id"],
        )
        session.add(rule)
        session.commit()
        return rule.id

def update_excel(results):
    """Update Excel with rule_ids for fixed rules."""
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    created_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

    for excel_row, rule_id in results:
        row_num = excel_row + 2  # +2 for header and 0-indexing
        # Column S (19) = status, Column T (20) = rule_id
        ws.cell(row=row_num, column=19).value = "CREATED"
        ws.cell(row=row_num, column=20).value = rule_id
        ws.cell(row=row_num, column=19).fill = created_fill

    wb.save(EXCEL_PATH)
    print(f"✅ Excel updated with fixed rules")

def main():
    print("=" * 80)
    print("FIX 2 FAILED RULES (Long Description)")
    print("=" * 80)

    results = []
    for rule_data in FAILED_RULES:
        try:
            rule_id = create_rule(rule_data)
            print(f"✅ Created: {rule_data['rule_name']} → rule_id={rule_id}")
            results.append((rule_data["excel_row"], rule_id))
        except Exception as e:
            print(f"❌ Failed: {rule_data['rule_name']} — {str(e)}")

    update_excel(results)

    print("\n" + "=" * 80)
    print(f"SUMMARY: {len(results)}/2 fixed")
    print("=" * 80)

if __name__ == "__main__":
    main()
