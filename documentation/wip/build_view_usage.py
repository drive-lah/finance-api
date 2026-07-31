#!/usr/bin/env python3
"""VIEW_USAGE — every ClickHouse view: used in the NEW economic-events system? + Dr/Cr accounts.

'Used' = wired in view_map.py (VIEW_MAP or PAYOUT_LINE_VIEWS) to an active template /
importer lane, or a feeder of such a view. NOT stripe_sync (dead), NOT what was posted for H1.
"""
import re, json, subprocess, glob, csv
from collections import defaultdict
from src.services.economic_events.view_map import VIEW_MAP, PAYOUT_LINE_VIEWS
from src.database import get_session_factory
from sqlalchemy import text
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CH = 'http://54.169.212.254:8123/?user=clickhouse-server-drivelah&password=Drivelah2025'
def ch(q): return subprocess.run(['curl', '-s', CH, '--data-binary', q], capture_output=True, text=True).stdout
views = {o['name']: o['create_table_query'] for o in (json.loads(r) for r in ch(
    "SELECT name,create_table_query FROM system.tables WHERE database='default' AND engine='View' AND (name LIKE 'view_%' OR name LIKE 'v\\_%') FORMAT JSONEachRow").strip().split('\n') if r.strip())}
allviews = set(views)

s = get_session_factory()()
ent = {'SG': 2, 'AU': 3}
acct = {c: n for c, n in s.execute(text("SELECT code,name FROM finance_accounts"))}
def A(code): return f"{code} {acct.get(code, '?')}" if code else ""
tmpl = defaultdict(list)
for et, eid, dr, cr in s.execute(text("SELECT event_type,entity_id,debit_code,credit_code FROM finance_je_templates WHERE is_active")):
    tmpl[et].append((eid, dr, cr))
active_ets = {(eid, et) for et in tmpl for (eid, _, _) in tmpl[et]}
bankcodes = {r[0] for r in s.execute(text("SELECT coa_account_code FROM finance_bank_accounts WHERE coa_account_code IS NOT NULL"))}
contra = {r[0] for r in s.execute(text("SELECT DISTINCT contra_account_code FROM finance_categorization_rules WHERE contra_account_code IS NOT NULL"))}
manual = bankcodes | contra
s.close()

view_wire = defaultdict(set)
active_ee = set()
for (m, et), sp in VIEW_MAP.items():
    if (ent.get(m), et) in active_ets:
        active_ee.add(sp.view)
        for (eid, dr, cr) in tmpl.get(et, []):
            if eid == ent.get(m):
                view_wire[sp.view].add((dr, cr))
payout_lane = {sp.view for sp in PAYOUT_LINE_VIEWS.values()}

deps = defaultdict(set)
for v, sql in views.items():
    for o in allviews:
        if o != v and re.search(r'\b' + re.escape(o) + r'\b', sql):
            deps[v].add(o)
def clo(seeds):
    seen = set(); st = list(seeds)
    while st:
        for f in deps.get(st.pop(), ()):
            if f not in seen:
                seen.add(f); st.append(f)
    return seen
roots = active_ee | payout_lane
feeders = clo(roots) - roots
read_by = defaultdict(set)
for v, ds in deps.items():
    for d in ds:
        read_by[d].add(v)

rows = []
for v in sorted(allviews):
    if v in active_ee:
        combos = sorted(view_wire[v])
        drs = "; ".join(A(dr) for dr, cr in combos)
        crs = "; ".join(A(cr) for dr, cr in combos)
        notes = []
        for dr, cr in combos:
            if dr in manual: notes.append(f"Dr {dr} also takes manual bank txns")
            if cr in manual: notes.append(f"Cr {cr} also takes manual bank txns")
        rows.append([v, "YES", drs, crs, "; ".join(sorted(set(notes)))])
    elif v in payout_lane:
        rows.append([v, "YES (payout lane)", "1017/1019 Bank - Stripe (Platform)", "",
                     "Stripe payout importer (/import-payouts) — lands as bank txns on the Stripe Platform account; not a template JE"])
    elif v in feeders:
        parent = sorted(x for x in read_by.get(v, ()) if x in active_ee)[:1]
        rows.append([v, "YES (feeder)", "", "", f"feeds {parent[0] if parent else 'active view'} — no direct posting"])
    else:
        rows.append([v, "NO", "", "", ""])

with open('documentation/wip/VIEW_USAGE.csv', 'w', newline='') as f:
    w = csv.writer(f); w.writerow(["View", "Used in new system?", "Debit account", "Credit account", "Note"])
    for r in rows: w.writerow(r)

wb = Workbook(); ws = wb.active; ws.title = "View Usage"
HF = PatternFill("solid", fgColor="1F4E78"); TH = Side(style="thin", color="BFBFBF"); BD = Border(TH, TH, TH, TH)
hdr = ["View", "Used in new system?", "Debit account (code + name)", "Credit account (code + name)", "Note"]
ws.append(hdr)
for c in range(1, 6):
    cell = ws.cell(1, c); cell.font = Font(bold=True, color="FFFFFF"); cell.fill = HF
    cell.alignment = Alignment(wrap_text=True, vertical="center"); cell.border = BD
fill = {"YES": "C6EFCE", "YES (payout lane)": "C6EFCE", "YES (feeder)": "FFF2CC", "NO": "F8CBAD"}
for r in rows:
    ws.append(r); rr = ws.max_row
    for c in range(1, 6):
        cell = ws.cell(rr, c); cell.border = BD; cell.alignment = Alignment(wrap_text=True, vertical="top"); cell.font = Font(size=10)
    ws.cell(rr, 2).fill = PatternFill("solid", fgColor=fill.get(r[1], "FFFFFF"))
for i, w in enumerate([46, 18, 32, 32, 42], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:E{ws.max_row}"
wb.save('documentation/wip/VIEW_USAGE.xlsx')

from collections import Counter
c = Counter(r[1] for r in rows)
print("counts:", dict(c), "| total", len(rows))
print("deposit/verif views (check 1021/1022):")
for r in rows:
    if 'deposit' in r[0].lower() or 'verif' in r[0].lower():
        print("  ", r[0], "|", r[1], "| Dr", r[2], "| Cr", r[3])
print("saved VIEW_USAGE.csv + .xlsx")
